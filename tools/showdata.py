import openpyxl
import showxml
import os

def get_allep():
    xlsx_path = showxml.findxlpath()
    if(xlsx_path == ""):
        return ""
    else:
        xlsx_file = os.path.join(xlsx_path,"episodeList.xlsx")
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        all_ep = {}
        mr = sheet.max_row
        for row in sheet.iter_rows():
            for cell in row:
                if(cell.row>1 and cell.column==1):
                    epno = str(cell.value)
                if(cell.row>1 and cell.column==2):
                    all_ep[epno] = str(cell.value)
        wb_obj.close()        
        return(all_ep)
        
def get_allcha():
    xlsx_path = showxml.findxlpath()
    if(xlsx_path == ""):
        return ""
    else:
        xlsx_file = os.path.join(xlsx_path,"chaList.xlsx").replace("\\","/")
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheetnames = wb_obj.sheetnames
        all_ep = []
        for i in range (0,3):
            wb_obj.active = i
            sheet = wb_obj.active
            for row in sheet.iter_rows():
                abc = [sheetnames[i]]
                for cell in row:
                    if(cell.row>1 and cell.column==1 and cell.value!=None):
                        abc.append(cell.value)
                    if(cell.row>1 and cell.column==2 and cell.value!=None):
                        abc.append(cell.value)
                    if(len(abc) == 3):
                        all_ep.append(abc)
        wb_obj.close()
        return(all_ep)
        
def get_allprop(episode):
    xlsx_path = showxml.findxlpath()
    if(xlsx_path == ""):
        return ""
    else:
        xlsx_file = os.path.join(xlsx_path,(episode+".xlsx"))
        chk_fe = os.path.isfile(xlsx_file)
        if(chk_fe == False):
            return ""
        else:
            wb_obj = openpyxl.load_workbook(xlsx_file)
            sheet = wb_obj.active
            all_prop = []
            mr = sheet.max_row
            for row in sheet.iter_rows():
                for cell in row:
                    if(cell.row>1 and cell.column==1):
                        all_prop.append(cell.value)
            wb_obj.close()        
            return(all_prop)

def get_allbg():
    xlsx_path = showxml.findxlpath()
    if (xlsx_path == ""):
        return ""
    else:
        xlsx_file = os.path.join(xlsx_path, "bgList.xlsx")
        chk_fe = os.path.isfile(xlsx_file)
        if (chk_fe == False):
            return ""
        else:
            wb_obj = openpyxl.load_workbook(xlsx_file)
            sheet = wb_obj.active
            all_bgs = []
            for row in sheet.iter_rows():
                bg_list = []
                for cell in row:
                    if (cell.row > 1 and cell.column == 1):
                        if(cell.value != None):
                            bg_list.append(cell.value)
                    if (cell.row > 1 and cell.column == 2):
                        if (cell.value != None):
                            bg_list.append(cell.value)
                        elif(len(bg_list) != 0):
                            bg_list.append("None")
                    if (cell.row > 1 and cell.column == 3):
                        if (cell.value != None):
                            bg_list.append(cell.value)
                        elif(len(bg_list) != 0):
                            bg_list.append("None")
                if(len(bg_list) != 0):
                    all_bgs.append(bg_list)
            wb_obj.close()
            return (all_bgs)

if __name__ == "__main__":
    print(get_allbg())